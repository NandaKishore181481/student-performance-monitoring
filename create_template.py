"""
create_template.py
------------------
Generates a college_data_template.xlsx file that you can fill in with your real 
college data. After filling, run import_data.py to load it into the system.

Usage:
    python create_template.py
"""

import os
import sys

# Fix Windows terminal encoding for emoji/unicode output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "college_data_template.xlsx")
# ─────────────────────────────────────────────

# Color palette
DARK_BLUE   = "1F3864"
LIGHT_BLUE  = "BDD7EE"
HEADER_GOLD = "FFD966"
LIGHT_YELLOW= "FFF2CC"
LIGHT_GREEN = "E2EFDA"
LIGHT_ORANGE= "FCE4D6"
WHITE       = "FFFFFF"
GRAY        = "D9D9D9"

def header_font(bold=True, size=11, color="FFFFFF"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def header_fill(color=DARK_BLUE):
    return PatternFill("solid", fgColor=color)

def cell_fill(color=WHITE):
    return PatternFill("solid", fgColor=color)

def thin_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, bg=DARK_BLUE, fg="FFFFFF"):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font  = header_font(color=fg)
        cell.fill  = header_fill(bg)
        cell.border = thin_border()
        cell.alignment = center()

def style_data_row(ws, row, cols, bg=WHITE):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = cell_font()
        cell.fill = cell_fill(bg)
        cell.border = thin_border()
        cell.alignment = center()

def add_note(ws, row, col, text):
    """Add a small note row below headers."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font  = Font(name="Calibri", italic=True, size=9, color="595959")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════════
def make_instructions_sheet(wb):
    ws = wb.create_sheet("📋 Instructions", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 90

    title_fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_font = Font(name="Calibri", bold=True, size=16, color="FFD966")

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = "  🎓 College Data Import Template — Instructions"
    ws["A1"].font  = title_font
    ws["A1"].fill  = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    instructions = [
        ("", ""),
        ("📌 OVERVIEW", ""),
        ("", "This workbook contains 4 sheets you must fill in before importing:"),
        ("", "  1. 'Departments'   — list your departments (e.g. CS, ECE, Mech)"),
        ("", "  2. 'Staff'         — HOD + Faculty names, emails, usernames"),
        ("", "  3. 'Students'      — Student details, roll number, section, year"),
        ("", "  4. 'Marks'         — Academic marks for each student per subject"),
        ("", "  5. 'Attendance'    — Per-day/per-subject attendance records"),
        ("", ""),
        ("📌 COLUMN RULES", ""),
        ("", "• username    : Unique login ID. E.g. cs_john2024  (no spaces, no @)"),
        ("", "• password    : Default login password for that user"),
        ("", "• roll_number : Must be unique across ALL students. E.g. CS2024001"),
        ("", "• department  : Must match exactly what you write in Departments sheet"),
        ("", "• year        : Must be 1, 2, 3, or 4"),
        ("", "• section     : Section letter or name. E.g. A, B, or A1"),
        ("", "• attendance  : A number between 0 and 100 (percentage)"),
        ("", "• internal    : Marks out of 30"),
        ("", "• assignment  : Marks out of 20"),
        ("", "• exam        : Marks out of 50  (leave blank if not yet conducted)"),
        ("", ""),
        ("📌 ATTENDANCE RULES", ""),
        ("", "• date        : DD-MM-YYYY format (e.g. 15-06-2024)"),
        ("", "• subject     : Subject name — must match the name used in Marks sheet"),
        ("", "• status      : Must be 'Present', 'Absent', or 'Late'"),
        ("", "• attendance_pct will be AUTO-CALCULATED from the Attendance sheet"),
        ("", "  (Leave attendance_pct blank in Students sheet if filling Attendance sheet)"),
        ("", ""),
        ("📌 HOW TO IMPORT", ""),
        ("", "1. Fill in all 5 sheets completely."),
        ("", "2. Save this file in the project folder as 'college_data_template.xlsx'"),
        ("", "3. Stop the running server (Ctrl+C in the terminal)."),
        ("", "4. Open a terminal in the project folder and run:"),
        ("", "        python import_data.py"),
        ("", "5. This will WIPE the demo data and load your college data."),
        ("", "6. Restart the server:  python run.py --api"),
        ("", "7. Restart dashboard:   python run.py --dashboard"),
        ("", ""),
        ("⚠️  WARNING", "Do NOT delete the header rows (row 1) in any sheet. Only fill from row 2 onward."),
        ("", "The importer will skip blank rows automatically."),
    ]

    for i, (key, val) in enumerate(instructions, start=2):
        ws.row_dimensions[i].height = 18
        kc = ws.cell(row=i, column=1, value=key)
        vc = ws.cell(row=i, column=2, value=val)
        if key and key.startswith("📌"):
            kc.font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            vc.font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif key and key.startswith("⚠️"):
            kc.font = Font(name="Calibri", bold=True, size=10, color="C00000")
            vc.font = Font(name="Calibri", bold=True, size=10, color="C00000")
        else:
            kc.font = Font(name="Calibri", size=10)
            vc.font = Font(name="Calibri", size=10)


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: DEPARTMENTS
# ═══════════════════════════════════════════════════════════════════
def make_departments_sheet(wb):
    ws = wb.create_sheet("🏛 Departments")
    ws.sheet_view.showGridLines = False

    headers = ["department_code", "department_name", "hod_username"]
    notes   = ["Short code (e.g. CS)", "Full name (e.g. Computer Science)", "Username of HOD from Staff sheet"]
    widths  = [20, 35, 25]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), bg=DARK_BLUE)

    ws.row_dimensions[2].height = 18
    for i, note in enumerate(notes, 1):
        add_note(ws, 2, i, note)

    # Sample rows
    samples = [
        ["CS", "Computer Science", "hod_cs"],
        ["ECE", "Electronics and Communication", "hod_ece"],
        ["MECH", "Mechanical Engineering", "hod_mech"],
    ]
    for r, row in enumerate(samples, start=3):
        ws.row_dimensions[r].height = 18
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), bg=LIGHT_GREEN if r % 2 == 0 else WHITE)


# ═══════════════════════════════════════════════════════════════════
# SHEET 3: STAFF (HOD + Faculty)
# ═══════════════════════════════════════════════════════════════════
def make_staff_sheet(wb):
    ws = wb.create_sheet("👥 Staff")
    ws.sheet_view.showGridLines = False

    headers = [
        "username", "password", "full_name", "email",
        "phone", "role", "department_code", "subjects_taught"
    ]
    notes = [
        "Unique login (no spaces)",
        "Default password",
        "Full name",
        "Email address",
        "Phone (optional)",
        "HOD or Faculty",
        "From Departments sheet",
        "Comma separated (Faculty only)"
    ]
    widths = [20, 15, 28, 35, 18, 10, 18, 40]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), bg=DARK_BLUE)

    ws.row_dimensions[2].height = 18
    for i, note in enumerate(notes, 1):
        add_note(ws, 2, i, note)

    # Role validation
    dv = DataValidation(type="list", formula1='"HOD,Faculty"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = "F3:F1000"

    # Sample rows
    samples = [
        ["hod_cs",      "hod123",  "Dr. A. Kumar",        "hod.cs@college.edu",    "9876543210", "HOD",     "CS",   ""],
        ["hod_ece",     "hod123",  "Dr. P. Rao",           "hod.ece@college.edu",   "9876543211", "HOD",     "ECE",  ""],
        ["fac_cs_math", "fac123",  "Prof. S. Sharma",      "s.sharma@college.edu",  "9876500001", "Faculty", "CS",   "Mathematics, Physics"],
        ["fac_cs_cs",   "fac123",  "Prof. R. Nair",        "r.nair@college.edu",    "9876500002", "Faculty", "CS",   "Computer Science, Data Structures"],
        ["fac_ece_elec","fac123",  "Prof. M. Iyer",        "m.iyer@college.edu",    "9876500003", "Faculty", "ECE",  "Electronics, Circuits"],
    ]
    for r, row in enumerate(samples, start=3):
        ws.row_dimensions[r].height = 18
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), bg=LIGHT_BLUE if r % 2 == 0 else WHITE)


# ═══════════════════════════════════════════════════════════════════
# SHEET 4: STUDENTS
# ═══════════════════════════════════════════════════════════════════
def make_students_sheet(wb):
    ws = wb.create_sheet("🎓 Students")
    ws.sheet_view.showGridLines = False

    headers = [
        "username", "password", "full_name", "email", "phone",
        "roll_number", "department_code", "year", "section",
        "attendance_pct",
        "parent_name", "parent_email", "parent_phone"
    ]
    notes = [
        "Login ID", "Password", "Full name", "Student email", "Student phone",
        "Unique roll no.", "From Departments", "1/2/3/4", "A/B/C",
        "0-100 (%)",
        "Parent full name", "Parent email", "Parent phone"
    ]
    widths = [18, 12, 25, 32, 14, 15, 18, 8, 10, 15, 25, 32, 14]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), bg=DARK_BLUE)

    ws.row_dimensions[2].height = 18
    for i, note in enumerate(notes, 1):
        add_note(ws, 2, i, note)

    # Year validation
    dv_year = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    ws.add_data_validation(dv_year)
    dv_year.sqref = "H3:H5000"

    # Sample rows
    samples = [
        ["cs_john_2024", "student123", "John Raj",    "john.raj@student.edu",   "9000000001", "CS2024001", "CS",   1, "A", 85.0, "Mr. A Raj",   "a.raj@gmail.com",   "9100000001"],
        ["cs_priya_2024","student123", "Priya Kumar", "priya.k@student.edu",    "9000000002", "CS2024002", "CS",   1, "A", 91.5, "Mrs. K Kumar", "k.kumar@gmail.com", "9100000002"],
        ["cs_ravi_2023", "student123", "Ravi Shankar","ravi.s@student.edu",     "9000000003", "CS2023001", "CS",   2, "B", 72.0, "Mr. S Kumar", "s.kumar@gmail.com", "9100000003"],
        ["ece_meena_2024","student123","Meena Devi",  "meena.d@student.edu",    "9000000004", "EC2024001", "ECE",  1, "A", 67.5, "Mr. D Iyer",  "d.iyer@gmail.com",  "9100000004"],
    ]
    for r, row in enumerate(samples, start=3):
        ws.row_dimensions[r].height = 18
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), bg=LIGHT_YELLOW if r % 2 == 0 else WHITE)


# ═══════════════════════════════════════════════════════════════════
# SHEET 5: MARKS
# ═══════════════════════════════════════════════════════════════════
def make_marks_sheet(wb):
    ws = wb.create_sheet("📊 Marks")
    ws.sheet_view.showGridLines = False

    headers = [
        "roll_number", "subject",
        "internal_marks",  # out of 30
        "assignment_scores",  # out of 20
        "exam_marks"  # out of 50 (can be blank)
    ]
    notes = [
        "Must match Students sheet",
        "Subject name (e.g. Mathematics)",
        "Out of 30",
        "Out of 20",
        "Out of 50 (leave blank if not done)"
    ]
    widths = [18, 35, 18, 20, 18]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), bg=DARK_BLUE)

    ws.row_dimensions[2].height = 18
    for i, note in enumerate(notes, 1):
        add_note(ws, 2, i, note)

    # Sample marks
    samples = [
        ["CS2024001", "Mathematics",       25, 17, 42],
        ["CS2024001", "Physics",           22, 15, 38],
        ["CS2024001", "Computer Science",  28, 19, 46],
        ["CS2024001", "English",           20, 14, 33],
        ["CS2024002", "Mathematics",       18, 12, 28],
        ["CS2024002", "Physics",           16, 11, 24],
        ["CS2024002", "Computer Science",  14, 10, 22],
        ["CS2024002", "English",           17, 12, 26],
        ["CS2023001", "Mathematics",       27, 18, 44],
        ["CS2023001", "Physics",           26, 17, 43],
        ["CS2023001", "Computer Science",  29, 19, 48],
        ["CS2023001", "English",           24, 16, 40],
        ["EC2024001", "Electronics",       21, 14, 35],
        ["EC2024001", "Mathematics",       19, 13, 30],
        ["EC2024001", "Physics",           23, 15, 37],
        ["EC2024001", "English",           20, 13, 32],
    ]
    for r, row in enumerate(samples, start=3):
        ws.row_dimensions[r].height = 18
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), bg=LIGHT_ORANGE if r % 2 == 0 else WHITE)


# ═══════════════════════════════════════════════════════════════════
# SHEET 6: ATTENDANCE
# ═══════════════════════════════════════════════════════════════════
LIGHT_RED   = "FCE4D6"
LIGHT_PURPLE= "EAD1DC"

def make_attendance_sheet(wb):
    ws = wb.create_sheet("📅 Attendance")
    ws.sheet_view.showGridLines = False

    headers = [
        "roll_number",
        "date",          # DD-MM-YYYY
        "subject",
        "status"         # Present / Absent / Late
    ]
    notes = [
        "Must match Students sheet roll_number",
        "Format: DD-MM-YYYY (e.g. 15-06-2024)",
        "Subject name (e.g. Mathematics)",
        "Present / Absent / Late"
    ]
    widths = [20, 18, 35, 15]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), bg="5B2C6F", fg="FFFFFF")

    ws.row_dimensions[2].height = 18
    for i, note in enumerate(notes, 1):
        add_note(ws, 2, i, note)

    # Status dropdown validation
    dv_status = DataValidation(
        type="list",
        formula1='"Present,Absent,Late"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = "D3:D50000"

    # Status cell colours
    PRESENT_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
    ABSENT_FILL  = PatternFill("solid", fgColor="FFC7CE")  # light red
    LATE_FILL    = PatternFill("solid", fgColor="FFEB9C")  # light yellow

    from datetime import date, timedelta
    today = date.today()
    subjects_cs  = ["Mathematics", "Physics", "Computer Science", "English"]
    subjects_ece = ["Electronics", "Mathematics", "Physics", "English"]

    # Sample attendance rows (last 5 working days)
    samples = []
    for offset in range(5, 0, -1):
        d = today - timedelta(days=offset)
        d_str = d.strftime("%d-%m-%Y")
        for subj in subjects_cs:
            samples.append(["CS2024001", d_str, subj, "Present"])
            samples.append(["CS2024002", d_str, subj, "Absent" if offset in (2, 4) else "Present"])
            samples.append(["CS2023001", d_str, subj, "Present"])
        for subj in subjects_ece:
            samples.append(["EC2024001", d_str, subj, "Late" if offset == 3 else "Present"])

    for r, row in enumerate(samples, start=3):
        ws.row_dimensions[r].height = 16
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font   = cell_font()
            cell.border = thin_border()
            cell.alignment = center()
            # colour-code the status cell
            if c == 4:
                if val == "Present":
                    cell.fill = PRESENT_FILL
                elif val == "Absent":
                    cell.fill = ABSENT_FILL
                else:
                    cell.fill = LATE_FILL
            else:
                cell.fill = cell_fill(WHITE)

    # Freeze the header rows
    ws.freeze_panes = "A3"

    # Auto-filter
    ws.auto_filter.ref = f"A1:D{len(samples) + 2}"


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    print("Creating Excel template...")
    make_instructions_sheet(wb)
    make_departments_sheet(wb)
    make_staff_sheet(wb)
    make_students_sheet(wb)
    make_marks_sheet(wb)
    make_attendance_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Template created: {OUTPUT_FILE}")
    print("\nNext steps:")
    print("  1. Open 'college_data_template.xlsx'")
    print("  2. Fill in your real college data in each sheet")
    print("  3. Save the file")
    print("  4. Run:  python import_data.py")
    print("     This will clear demo data and load your real data.\n")


if __name__ == "__main__":
    main()
