"""
import_data.py
-------------
Reads 'college_data_template.xlsx' and imports the data into the system database.
This will REPLACE all demo/sample data with your real college data.

Usage:
    python import_data.py                        # imports from college_data_template.xlsx
    python import_data.py --file my_data.xlsx    # imports from a custom file
    python import_data.py --preview              # shows what will be imported (no changes made)
"""

import os
import sys

# Fix Windows terminal encoding for emoji/unicode output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import argparse
from datetime import date, datetime

# ── Make sure src/ is importable ──────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR  = os.path.join(BASE_DIR, "src")
sys.path.insert(0, SRC_DIR)

try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("❌ Missing libraries. Run: pip install openpyxl pandas")
    sys.exit(1)

from database import (
    engine, SessionLocal, Base,
    User, StudentProfile, AcademicMarks,
    FacultyRemarks, Assignment, AlertLog,
    AttendanceRecord,
    hash_password
)

DEFAULT_FILE = os.path.join(BASE_DIR, "college_data_template.xlsx")


# ═══════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════

def read_sheet(wb, sheet_name):
    """Read a worksheet into a list of dicts, skipping blank rows and the notes row."""
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  Sheet '{sheet_name}' not found — skipping.")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    # Row index 1 (0-based) is the notes row — skip it
    data = []
    for row in rows[2:]:
        # Skip completely blank rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {}
        for h, val in zip(headers, row):
            record[h] = val
        data.append(record)
    return data


def s(val, default=""):
    """Safe string conversion."""
    if val is None:
        return default
    return str(val).strip()


def f(val, default=0.0):
    """Safe float conversion."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def validate_file(filepath):
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        print("   Run  python create_template.py  to generate the template first.")
        sys.exit(1)


# ═══════════════════════════════════════════════════════════════════
# VALIDATION & PREVIEW
# ═══════════════════════════════════════════════════════════════════

def validate_and_preview(filepath):
    """Validate the spreadsheet and show a summary."""
    print(f"\n📂 Reading: {filepath}\n")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    dept_rows      = read_sheet(wb, "🏛 Departments")
    staff_rows     = read_sheet(wb, "👥 Staff")
    student_rows   = read_sheet(wb, "🎓 Students")
    marks_rows     = read_sheet(wb, "📊 Marks")
    attendance_rows= read_sheet(wb, "📅 Attendance")

    errors = []

    # Collect department codes
    dept_codes = set()
    for d in dept_rows:
        code = s(d.get("department_code"))
        if code:
            dept_codes.add(code)

    # Collect staff usernames
    staff_usernames = set()
    for st in staff_rows:
        u = s(st.get("username"))
        if u:
            staff_usernames.add(u)

    # Collect student roll numbers
    roll_numbers = set()
    for sv in student_rows:
        rn = s(sv.get("roll_number"))
        if rn:
            if rn in roll_numbers:
                errors.append(f"Duplicate roll number: {rn}")
            roll_numbers.add(rn)
        dept = s(sv.get("department_code"))
        if dept and dept not in dept_codes:
            errors.append(f"Student {rn}: department '{dept}' not in Departments sheet")
        yr = sv.get("year")
        try:
            yr_int = int(yr)
            if yr_int not in [1, 2, 3, 4]:
                errors.append(f"Student {rn}: year must be 1-4, got {yr_int}")
        except (TypeError, ValueError):
            errors.append(f"Student {rn}: invalid year value '{yr}'")

    # Validate marks reference roll numbers
    marks_warnings = []
    for m in marks_rows:
        rn = s(m.get("roll_number"))
        if rn and rn not in roll_numbers:
            marks_warnings.append(f"Marks: roll_number '{rn}' not in Students (will be skipped)")
        internal = m.get("internal_marks")
        assign   = m.get("assignment_scores")
        try:
            if internal is not None and float(internal) > 30:
                errors.append(f"Marks {rn}: internal_marks {internal} > 30 (max allowed)")
            if assign is not None and float(assign) > 20:
                errors.append(f"Marks {rn}: assignment_scores {assign} > 20 (max allowed)")
            exam = m.get("exam_marks")
            if exam is not None and exam != "" and float(exam) > 50:
                errors.append(f"Marks {rn}: exam_marks {exam} > 50 (max allowed)")
        except (TypeError, ValueError):
            pass

    # Validate attendance rows
    valid_statuses = {"Present", "Absent", "Late"}
    att_warnings = []
    for a in attendance_rows:
        rn     = s(a.get("roll_number"))
        status = s(a.get("status", ""))
        date_raw = a.get("date")
        if rn and rn not in roll_numbers:
            att_warnings.append(f"Attendance: roll_number '{rn}' not in Students (will be skipped)")
        if status and status.capitalize() not in valid_statuses:
            errors.append(f"Attendance {rn}: status '{status}' must be Present/Absent/Late")
        # Validate date — accept datetime objects, date objects, or DD-MM-YYYY strings
        if date_raw is not None:
            if isinstance(date_raw, (datetime, date)):
                pass  # valid
            else:
                date_str = str(date_raw).strip()
                valid_date = False
                for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        datetime.strptime(date_str, fmt)
                        valid_date = True
                        break
                    except ValueError:
                        continue
                if not valid_date:
                    errors.append(f"Attendance {rn}: date '{date_str}' format not recognized")

    print("=" * 60)
    print("  📋  DATA PREVIEW")
    print("=" * 60)
    print(f"  Departments       : {len(dept_rows)}")
    print(f"  Staff members     : {len(staff_rows)}")
    print(f"  Students          : {len(student_rows)}")
    print(f"  Mark records      : {len(marks_rows)}")
    print(f"  Attendance records: {len(attendance_rows)}")
    print("=" * 60)

    # Show non-blocking warnings
    all_warnings = []
    if marks_warnings:
        unique_mw = list(dict.fromkeys(marks_warnings))
        all_warnings.extend(unique_mw)
    if att_warnings:
        unique_aw = list(dict.fromkeys(att_warnings))
        all_warnings.extend(unique_aw)
    
    if all_warnings:
        print(f"\n  ⚠️  {len(all_warnings)} warning(s) (rows will be skipped during import):")
        for w in all_warnings[:15]:
            print(f"     • {w}")
        if len(all_warnings) > 15:
            print(f"     ... and {len(all_warnings) - 15} more")
        print()

    if errors:
        print(f"\n  ❌ {len(errors)} validation error(s) found:\n")
        for e in errors:
            print(f"     • {e}")
        print()
        return False, wb, dept_rows, staff_rows, student_rows, marks_rows, attendance_rows

    print("\n  ✅ No validation errors found!\n")
    return True, wb, dept_rows, staff_rows, student_rows, marks_rows, attendance_rows



# ═══════════════════════════════════════════════════════════════════
# DATABASE WIPE
# ═══════════════════════════════════════════════════════════════════

def wipe_database(db):
    print("🗑️  Clearing existing data...")
    db.query(AlertLog).delete()
    db.query(Assignment).delete()
    db.query(FacultyRemarks).delete()
    db.query(AcademicMarks).delete()
    db.query(AttendanceRecord).delete()
    db.query(StudentProfile).delete()
    db.query(User).delete()
    db.commit()
    print("   ✅ Database cleared.\n")


# ═══════════════════════════════════════════════════════════════════
# IMPORT
# ═══════════════════════════════════════════════════════════════════

def import_data(dept_rows, staff_rows, student_rows, marks_rows, attendance_rows):
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        wipe_database(db)

        # ── Track created objects ───────────────────────────────
        dept_hod_map = {}   # dept_code -> hod_username (to assign later)
        username_user_map = {}  # username -> User object
        roll_profile_map = {}   # roll_number -> StudentProfile object

        # ── 1a. Auto-create HOD users from Departments sheet ──
        print("👥 Importing staff...")
        hod_count = 0
        fac_count = 0

        # Build set of staff usernames for lookup
        staff_usernames_set = {s(r.get("username", "")) for r in staff_rows}

        for dept in dept_rows:
            hod_uname = s(dept.get("hod_username", ""))
            dept_code = s(dept.get("department_code", ""))
            dept_name = s(dept.get("department_name", dept_code))
            if not hod_uname or not dept_code:
                continue
            # Only auto-create if HOD username is NOT already in staff sheet
            if hod_uname not in staff_usernames_set:
                hod_user = User(
                    username=hod_uname,
                    hashed_password=hash_password("hod123"),
                    role="HOD",
                    name=f"HOD - {dept_name}",
                    email=f"{hod_uname}@college.edu",
                    phone=""
                )
                db.add(hod_user)
                db.flush()
                username_user_map[hod_uname] = hod_user
                dept_hod_map[dept_code] = hod_user.id
                hod_count += 1

        # ── 1b. Create Faculty from Staff sheet ────────────────
        # Accept any role value (Professor, Assistant Prof, etc.) and map to "Faculty"
        # If someone is explicitly "HOD" in the role column, treat as HOD
        for row in staff_rows:
            username = s(row.get("username"))
            if not username:
                continue
            raw_role = s(row.get("role", "Faculty"))
            # Map role: only "HOD" stays as HOD, everything else becomes Faculty
            if raw_role.upper() == "HOD":
                role = "HOD"
            else:
                role = "Faculty"

            dept_code = s(row.get("department_code", ""))

            user = User(
                username=username,
                hashed_password=hash_password(s(row.get("password", "changeme123"))),
                role=role,
                name=s(row.get("full_name", username)),
                email=s(row.get("email", f"{username}@college.edu")),
                phone=s(row.get("phone", ""))
            )
            db.add(user)
            db.flush()
            username_user_map[username] = user

            if role == "HOD" and dept_code:
                dept_hod_map[dept_code] = user.id
                hod_count += 1
            else:
                fac_count += 1

        db.commit()
        print(f"   ✅ {hod_count} HODs, {fac_count} Faculty imported.\n")

        # ── 2. Create Parents + Students ───────────────────────
        print("🎓 Importing students...")
        student_count = 0
        for row in student_rows:
            username  = s(row.get("username"))
            roll      = s(row.get("roll_number"))
            if not username or not roll:
                continue

            # Create parent user first (if data provided)
            parent_id = None
            parent_name  = s(row.get("parent_name", ""))
            parent_email = s(row.get("parent_email", ""))
            parent_phone = s(row.get("parent_phone", ""))

            if parent_name and parent_email:
                parent_username = f"parent_{username}"
                parent_user = User(
                    username=parent_username,
                    hashed_password=hash_password("parent123"),
                    role="Parent",
                    name=parent_name,
                    email=parent_email,
                    phone=parent_phone
                )
                db.add(parent_user)
                db.flush()
                parent_id = parent_user.id
                username_user_map[parent_username] = parent_user

            # Create student user
            year    = int(f(row.get("year", 1)))
            section = s(row.get("section", "A"))
            dept    = s(row.get("department_code", ""))
            # class_section encodes dept + year + section for display
            class_section = f"{dept}-Y{year}{section}" if dept else f"Y{year}{section}"

            student_user = User(
                username=username,
                hashed_password=hash_password(s(row.get("password", "student123"))),
                role="Student",
                name=s(row.get("full_name", username)),
                email=s(row.get("email", f"{username}@student.edu")),
                phone=s(row.get("phone", ""))
            )
            db.add(student_user)
            db.flush()
            username_user_map[username] = student_user

            profile = StudentProfile(
                user_id=student_user.id,
                parent_id=parent_id,
                roll_number=roll,
                class_section=class_section,
                attendance_pct=f(row.get("attendance_pct", 75.0))
            )
            db.add(profile)
            db.flush()
            roll_profile_map[roll] = profile
            student_count += 1

        db.commit()
        print(f"   ✅ {student_count} students imported.\n")

        # ── 3. Import Marks ─────────────────────────────────────
        print("📊 Importing marks...")
        marks_count = 0
        skipped_marks = 0
        for row in marks_rows:
            roll    = s(row.get("roll_number"))
            subject = s(row.get("subject"))
            if not roll or not subject:
                continue

            profile = roll_profile_map.get(roll)
            if not profile:
                print(f"   ⚠️  Skipping marks: roll '{roll}' not found.")
                skipped_marks += 1
                continue

            exam_raw  = row.get("exam_marks")
            exam_val  = None
            if exam_raw is not None and str(exam_raw).strip() not in ("", "None"):
                try:
                    exam_val = float(exam_raw)
                except ValueError:
                    exam_val = None

            mark = AcademicMarks(
                student_id=profile.id,
                subject=subject,
                internal_marks=f(row.get("internal_marks", 0.0)),
                assignment_scores=f(row.get("assignment_scores", 0.0)),
                exam_marks=exam_val
            )
            db.add(mark)
            marks_count += 1

        db.commit()
        print(f"   ✅ {marks_count} mark records imported.")
        if skipped_marks:
            print(f"   ⚠️  {skipped_marks} mark records skipped (roll not found).\n")
        else:
            print()

        # ── 4. Import Attendance ────────────────────────────────
        att_count   = 0
        att_skipped = 0
        # Track per-student Percentage column values (if provided)
        # {roll: percentage_value}
        att_pct_from_sheet = {}
        # Also track computed totals as fallback
        att_totals  = {}

        if attendance_rows:
            print("📅 Importing attendance records...")
            for row in attendance_rows:
                roll    = s(row.get("roll_number"))
                date_raw = row.get("date")
                subject = s(row.get("subject", ""))
                status  = s(row.get("status", "Present"))

                if not roll or not subject:
                    continue

                profile = roll_profile_map.get(roll)
                if not profile:
                    att_skipped += 1
                    continue

                # Parse date — handle both datetime objects (from Excel) and strings
                parsed_date = None
                if isinstance(date_raw, datetime):
                    parsed_date = date_raw.date()
                elif isinstance(date_raw, date):
                    parsed_date = date_raw
                elif date_raw:
                    date_str = str(date_raw).strip()
                    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                        try:
                            parsed_date = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                if not parsed_date:
                    print(f"   ⚠️  Bad date '{date_raw}' for roll {roll} — skipping row.")
                    att_skipped += 1
                    continue

                # Normalise status
                status = s(status).capitalize()
                if status not in ("Present", "Absent", "Late"):
                    status = "Absent"

                rec = AttendanceRecord(
                    student_id=profile.id,
                    date=parsed_date,
                    subject=subject,
                    status=status
                )
                db.add(rec)
                att_count += 1

                # Capture Percentage column value (user-provided overall %)
                pct_val = row.get("Percentage") or row.get("percentage")
                if pct_val is not None and roll not in att_pct_from_sheet:
                    try:
                        att_pct_from_sheet[roll] = float(pct_val)
                    except (TypeError, ValueError):
                        pass

                # Also accumulate computed totals as fallback
                if roll not in att_totals:
                    att_totals[roll] = {"total": 0, "present": 0}
                att_totals[roll]["total"] += 1
                if status in ("Present", "Late"):
                    att_totals[roll]["present"] += 1

            db.commit()

            # Update attendance_pct — prefer the Percentage column, fallback to computed
            updated_pct = 0
            all_rolls = set(list(att_pct_from_sheet.keys()) + list(att_totals.keys()))
            for roll in all_rolls:
                profile = roll_profile_map.get(roll)
                if not profile:
                    continue
                if roll in att_pct_from_sheet:
                    profile.attendance_pct = att_pct_from_sheet[roll]
                elif roll in att_totals and att_totals[roll]["total"] > 0:
                    profile.attendance_pct = round(
                        (att_totals[roll]["present"] / att_totals[roll]["total"]) * 100, 2
                    )
                updated_pct += 1
            db.commit()

            print(f"   ✅ {att_count} attendance records imported.")
            if att_pct_from_sheet:
                print(f"   ✅ {len(att_pct_from_sheet)} students' attendance_pct set from Percentage column.")
            print(f"   ✅ {updated_pct} students' attendance_pct updated total.")
            if att_skipped:
                print(f"   ⚠️  {att_skipped} attendance rows skipped.")
            print()
        else:
            print("📅 No Attendance sheet data found — using attendance_pct from Students sheet.\n")

        print("=" * 60)
        print("  🎉  IMPORT COMPLETE!")
        print("=" * 60)
        print(f"  Staff       : {hod_count + fac_count}")
        print(f"  Students    : {student_count}")
        print(f"  Marks       : {marks_count}")
        print(f"  Attendance  : {att_count}")
        print("=" * 60)
        print("\nDefault Passwords:")
        print("  HOD      : (as set in Staff sheet)")
        print("  Faculty  : (as set in Staff sheet)")
        print("  Students : (as set in Students sheet)")
        print("  Parents  : parent123 (auto-assigned)")
        print()
        print("Restart the server to see your data:")
        print("  python run.py --api")
        print("  python run.py --dashboard\n")

    except Exception as e:
        db.rollback()
        print(f"\n❌ Import failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Import college data from Excel into the Student Monitoring System"
    )
    parser.add_argument(
        "--file", default=DEFAULT_FILE,
        help="Path to the Excel file (default: college_data_template.xlsx)"
    )
    parser.add_argument(
        "--preview", action="store_true",
        help="Only validate and preview — do NOT import"
    )
    args = parser.parse_args()

    validate_file(args.file)
    ok, wb, dept_rows, staff_rows, student_rows, marks_rows, attendance_rows = validate_and_preview(args.file)

    if args.preview:
        print("ℹ️  Preview mode — no changes made.")
        return

    if not ok:
        print("❌ Fix the errors above, then run import again.")
        sys.exit(1)

    confirm = input("⚠️  This will WIPE all existing data. Continue? [y/N]: ").strip().lower()
    if confirm != "y":
        print("Aborted. No changes made.")
        sys.exit(0)

    import_data(dept_rows, staff_rows, student_rows, marks_rows, attendance_rows)


if __name__ == "__main__":
    main()
